import openpyxl
import re
# Constants
TEMPLATE_FN = 'template.xlsx'
DEFAULT_OUTPUT_FILE_NAME = 'new.xlsx'
MILLION_PAT = r"^\d+(,\d\d\d)+$"
FLOAT_PAT = r"^\d+\.\d+$"


def str_to_val(num_str):
	if type(num_str) is not str:
		return num_str
	if num_str.isnumeric():
		return int(num_str)
	if re.match(MILLION_PAT, num_str) is not None:
		return(int(num_str.replace(",", "")))
	if re.match(FLOAT_PAT, num_str) is not None:
		return(float(num_str))
	return num_str


class ExcelHandler(object):
	def __init__(self, output_file=DEFAULT_OUTPUT_FILE_NAME, table_row=1, table_col=1):
		self.output_file = output_file
		self.table_row = table_row
		self.current_row_in_table = 1  # Row after head of table
		self.table_col = table_col
		self.book = openpyxl.load_workbook(TEMPLATE_FN)
		self.sheet = self.book.active
		
	def get_table_head(self, num_of_cells):
		table_head = []
		for i in range(num_of_cells):
			table_head.append(self.sheet.cell(row=self.table_row, column=self.table_col + i).value)
		return table_head
		
	def insert_list_to_row(self, data_list, table_row=None):
		if table_row is None:
			table_row = self.current_row_in_table
		for col in range(len(data_list)):
			self.write_cell_table(table_row, col, data_list[col])
		self.current_row_in_table = table_row + 1
	
	def write_cell_table(self, row, col, val_str):
		val = str_to_val(val_str)
		self.sheet.cell(row=row + self.table_row, column=col + self.table_col).value = val
	
	def save(self):
		self.book.save(self.output_file)


if __name__ == '__main__':
	print("in main")
	eh = ExcelHandler("zvi.xlsx")
	print(eh.get_table_head(9))
	eh.insert_list_to_row([1, 2, 3])
	eh.save()
